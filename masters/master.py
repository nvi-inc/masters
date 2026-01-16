import tempfile
import re
import json

from operator import itemgetter
from datetime import datetime, date, time, timezone, timedelta
from collections import OrderedDict
from pathlib import Path
from string import Formatter
from typing import Dict, Tuple, List, Set, Optional, Any

import toml
from openpyxl.cell.read_only import EmptyCell
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import load_workbook

from masters import app, Base, get_file_name


class MasterFile(Formatter):
    """
    Formatter for master or intensive files
    """
    def __init__(self, code: str, year: str, version: Dict[str, str], fields: Dict[str or int, str]) -> None:
        """
        Initializes formatter for specific master file
        :param code: master ot master-media file
        :param year: Year in string of file
        :param version: Dictionary of file version for master and media
        :param fields: Dictionary formats for each field
        """
        super().__init__()

        self.path = Path(tempfile.gettempdir(), get_file_name(code, 'txt', year))
        self.header, self.separator, self.footer, self.line_format = self._build(code, year, version, fields)
        self.key, self.month = None, ''

    def __enter__(self):
        self.file = open(self.path, 'w+')
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.file.close()

    # Generate header, separation line, footer and format for data line
    @staticmethod
    def _build(code: str, year: str, version: Dict[str, str], fields: Dict[str or int, str]) -> Tuple[str, str, str]:
        """
        Build header
        :param code: master ot master-media file
        :param year: Year in string of file
        :param version: Dictionary of file version for master and media
        :param fields: Dictionary formats for each field
        :return: Header, line separator, format for session lines
        """
        version = version[code] if code in version else version['master']
        header = app.config[code]['header'].format(version=version, year=year,
                                                   updated=datetime.now(tz=timezone.utc).strftime('%B %d, %Y'),
                                                   initials=app.config['initials'])
        # Find the longest line to generate separator
        line_length = max([len(line) for line in header.splitlines()])
        separator = f"{'-' * line_length}\n"
        # Build format for this formatter
        columns = OrderedDict(**app.config['master']['format'])
        for key, value in app.config[code]['format'].items():
            columns[key] = value
        line_format = [fmt.replace('{', '{{{}:'.format(code)) for code, fmt in columns.items()]
        return header, separator, f"{separator}{version}\n", f"|{'|'.join(line_format)}|\n"

    # Write line
    def write(self, ses: Dict[str, Any]) -> None:
        """
        Control writing of session information into file
        :param ses:
        :return: None
        """
        if (month := ses['START'].strftime('%b')) != self.month:
            self.month = month
            self.file.write(self.separator)
        self.file.write(self.format(self.line_format, **ses))

    def write_header(self) -> None:
        """
        Control wrinting of header into file
        :return: None
        """
        self.file.write(self.header)

    # Write footer
    def write_footer(self):
        self.file.write(self.footer)

    # get the field name for the next format_field request
    def get_field(self, field_name, args, kwargs):
        self.key = field_name
        return super().get_field(field_name, args, kwargs)

    def format_field(self, value: Any, format_spec: str) -> str:
        """
        Control formatting of value
        :param value: Input value
        :param format_spec: Format for this value
        :return: Formatted value
        """
        if self.key in ['DATE', 'TIME']:
            value = value.strftime(format_spec).upper()
            format_spec = ''
        elif self.key == 'DUR' and format_spec == '%H:%M':
            hours, remainder = divmod(value, 1)
            value = f'{int(hours):2d}:{int(remainder*60):02d}' if value > 0 else '     '
            format_spec = ''
        elif self.key == 'STATUS':
            if isinstance(value, datetime):
                value = value.strftime(format_spec).upper()
                format_spec = ''
            else:
                format_spec = '<{:d}s'.format(len(date.today().strftime(format_spec).upper()))
        elif self.key in ['PF', 'MK4NUM'] and not isinstance(value, (int, float)):
            format_spec = '<{:d}s'.format(int(re.sub(r"\D", ' ', format_spec).strip().split()[0]))
        elif self.key == 'DELAY' and not isinstance(value, (int, float)):
            format_spec = format_spec.replace('d', 's')

        value = ' ' if value is None else value
        return format(value, format_spec)

    def get_value(self, key: str or int, args: List[Any], kwargs: Dict[str, Any]):
        """
        Overwritten function to control the selection of value
        :param key: Key or column of session information
        :param args: List of positional arguments if
        :param kwargs: Dictionary of keyword arguments
        :return: requested field value
        """
        self.key = key
        return kwargs[key] if isinstance(key, str) else super().get_value(key, args, kwargs)


class XLMaster(Base):
    """
    Class to process master Excel (xlsx) files
    """

    def __init__(self, path: Path) -> None:
        """
        Initialize class
        :param path: Path of the master xlsx file
        """
        super().__init__()
        # Extract information from filename
        self.path, self.name, self.ext = path, path.name, path.suffix
        self.year = re.sub(r"\D", ' ', self.name).strip()

        self.type, self.sessions, self.codes = None, [], set()

        self.folder = app.config['folder']
        self.version = {'master': 'UNKNOWN', 'media': 'UNKNOWN'}

        self.format_path, self.valid_codes = self.get_field_codes()
        self.ns_codes_path, self.ns_codes = self.get_ns_codes()
        self.media_key_path, self.media_sizes = self.get_media_keys()
        self.today = date.today()
        self.constrains = app.config['master']['constrains']
        self.fields = OrderedDict()
        self.old_code_constrain, self.fs10 = self.get_fs_10_stations()
        self.get_session_type = self.get_session_type_dict().get

    def get_file_path(self, code_name: str):
        """
        Get full file path using the code name
        :param code_name: master,
        :return: Path of file
        """
        # Check if code_name in configuration file
        name = app.config.get(code_name, f'{code_name}.txt')
        name = Path(name['path']).name if isinstance(name, dict) else name
        if (path := Path(self.folder, name)).exists():
            return path
        self.exit(error=f'{path} does not exist.')

    def get_field_codes(self) -> Tuple[Path, Dict[str, str]]:
        """
        Read master-format file and extract valid codes for specific fiels
        :return: Path of file and Dictionary of valid codes
        """
        path = self.get_file_path('master-format')
        # Read header and content
        with open(path) as file:
            header, content = file.readline(), file.read()
        # Test version
        if not (version := re.search(r'(## .*)', header)):
            self.exit(error=f'{path} is not a valid master-format file')
        self.version['master'] = version.group(1)
        # Load accepted data code for specific fields
        valid_codes, tag = {}, re.MULTILINE | re.DOTALL
        for field in re.findall(r'^\s*(\w*) CODES', content, tag):
            pattern = fr'^\s*{field} CODES(.*)^\s*end {field} CODES'
            valid_codes[field] = [l.split()[0] for l in re.findall(pattern, content, tag)[0].splitlines() if l.strip()]
        return path, valid_codes

    def get_ns_codes(self) -> Tuple[Path, List[str]]:
        """
        Read ns-codes file and extract station codes
        :return: Path of file and list of station codes
        """
        path = self.get_file_path('ns-codes')
        with open(path) as file:
            content = file.read()
        tag = re.MULTILINE | re.DOTALL
        ns_codes = [a[0] for a in re.findall(r'^ (\w{2}) (.{8})', content, tag) if a[1] != '--------']
        return path, ns_codes

    def get_media_keys(self) -> Tuple[Path, Set[str]]:
        """
        Read media-key file and extract disk size codes
        :return: Path file and set of accepted media sizes
        """
        path = self.get_file_path('media-key')
        with open(path) as file:
            tag = re.MULTILINE | re.DOTALL
            with open(path) as file:
                content = re.search(r'type of media(.*)', file.read(), tag).group(1)
            sizes = [a for a in re.findall(r'^\s+([a-zA-Z]) =', content, tag)]
        return path, set(sizes)

    def validate_station_info(self, ses: Dict[str, Any], cell: Cell, hdr: str) -> Tuple[str, str]:
        """
        Validate station information
        :param ses: Session information
        :param cell: Cell of sheet
        :param hdr: Name of column
        :return: Station code and value of cell
        """
        has_media = self.type == 'master'
        try:
            if not cell.value or not cell.value.strip():
                return '  ', '    '
        except AttributeError:
            print(cell.row, cell.column)
            return '  ', '    '
        value = cell.value  #.replace('-', '').strip()
        sta = value[0:2]
        if has_media and (not value[2].isdigit() or value[3] not in self.media_sizes):
            self.add_error(ses, f'invalid information [{value}] in column {hdr}')
        if sta not in self.ns_codes:
            self.add_error(ses, f'invalid station code [{sta}] in column {hdr}')

        return sta, value

    @staticmethod
    def format_list(stations: List[str], n: int) -> str:
        """
        Group by participating and non-participating stations and sort each groups
        :param stations: List of stations
        :param n: Size of string in station value
        :return: Formatted information ready for master text file
        """

        def clean(item):
            return ''.join(sorted(list(re.findall('.' * n, item.strip()))))

        groups = [clean(grp) for grp in ''.join([sta[:n] for sta in stations]).split()]
        return f" -{groups[0]}" if stations[0].strip() == '' else ' -'.join(groups)

    # Validate each session
    def validate_session(self, ses: Dict[str, Any], stations: List[str]) -> Dict[str, Any]:
        """
        Validate session information
        :param ses: Session data
        :param stations: List of stations
        :return: Validated session
        """
        if ses['CODE'].lower() in self.codes:
            self.add_error(ses , 'duplicate session name')
        self.codes.add(ses['CODE'].lower())
        for code, constrain in self.constrains.items():
            if len(ses[code]) > constrain:
                self.add_error(ses, f'{code} {ses[code]} has more than {constrain:d} characters')
        # Check that DATE, TIME and DOY are valid
        if not isinstance(ses['DATE'], datetime):
            self.add_error(ses, f'invalid DATE {str(ses["DATE"])}')
        elif ses['DATE'].strftime('%Y') != self.year:
            self.add_error(ses, f'invalid DATE {ses["DATE"].strftime("%Y-%m-%d")}')
        if not isinstance(ses['TIME'], time):
            self.add_error(ses, f'invalid TIME {str(ses["TIME"])}')
        if isinstance(ses['DATE'], date) and isinstance(ses['TIME'], time):
            ses['START'] = datetime.combine(ses['DATE'].date(), ses['TIME'])
            ses['DATE'] = ses['DATE'].date()
        if len(ses['CODE']) > self.old_code_constrain:
            if not_fs10 := [code for code in stations if code.strip() and code not in self.fs10]:
                self.add_error(ses, f'CODE too long for [{",".join(not_fs10)}]! Maximum is {self.old_code_constrain}')

        # Validate SKED, CORR, SUBM codes
        for code, items in self.valid_codes.items():
            if code not in ['STATUS', 'DBC'] and ses[code] not in items:
                self.add_error(ses, f'invalid {code} code {ses[code]}')

        # Extract EXPERIMENT name from master-type-map file when sessions before 2024
        if ses['DATE'].year < 2024:
            ses['EXPERIMENT'] = self.get_session_type(ses['CODE'].strip().casefold(), None)
            if not ses['EXPERIMENT']:
                self.add_error(ses, f"session {ses['CODE'].strip().casefold()} not found in master-type-map.json file")
        ses['CODE'] = ses['CODE'].lower()
        #elif ses['DOY'] is not None and int(ses['DATE'].strftime('%j')) != int(ses['DOY']):
        #    self.add_error(ses, f'invalid DOY {str(ses["DOY"])}')
        ses['DOY'] = int(ses['DATE'].strftime('%j'))
        if self.type == 'master' and self.year > '1997':  # Old master files has blank STATUS
            # Validate status codes if master schedule
            if not ses['STATUS']:
                if self.debug and ses['DATE'] <= self.today:
                    ses['STATUS'] = 'Wt_med'
                elif isinstance(ses['DATE'], date) and ses['DATE'] <= self.today:
                    self.add_error(ses, 'STATUS code is blank!', debug=self.debug)
            elif not isinstance(ses['STATUS'], (datetime, date)) and ses['STATUS'] not in self.valid_codes['STATUS']:
                self.add_error(ses, f'STATUS code {ses["STATUS"]} is not valid', debug=self.debug)

        return ses

    @staticmethod
    def read_header(row):
        fields = OrderedDict()
        for cell in row:
            if not isinstance(cell, EmptyCell) and isinstance(cell.value, str):
                fields[cell.column] = txt = cell.value.strip()
                #if txt.upper() in ('SUBM', 'MK4NUM'):
                #    break
        return fields

    def read_master(self, sheet: Worksheet) -> bool:
        """
        Read master sessions stored in Excel worksheet
        :param sheet: Active sheet
        :return: Success
        """
        # Read fields from first row
        rows = sheet.iter_rows()
        self.fields = self.read_header(next(rows))
        if 'DELAY' not in self.fields.values():
            self.fields[1000] = 'DELAY'

        # Read data
        sessions = []
        for row in rows:
            ses, master, media = {}, [], []
            for cell in row:
                if isinstance(cell, EmptyCell):
                    continue
                if 'row' not in ses:
                    ses['row'] = cell.row

                if field := self.fields.get(cell.column, None):
                    ses[field] = cell.value
                    if field.startswith('Stat'):
                        sta, size = self.validate_station_info(ses, cell, field)
                        media.append(size)
                        master.append(sta)

            if not ses.get('DATE', None):
                continue

            # Validate this session
            ses = self.validate_session(ses, master)
            # Format station codes and media
            ses['master'] = self.format_list(master, 2)
            ses['media'] = self.format_list(media, 4)

            if 'DELAY' not in ses:
                ses['DELAY'] = 0
            if self.year < '2003':
                ses['DELAY'] = ''
            else:  # Compute delay from status or today.
                end = ses['START'].astimezone(timezone.utc) + timedelta(seconds=ses['DUR'])
                if isinstance(ses['STATUS'], (datetime, date)):
                    ses['DELAY'] = (ses['STATUS'].astimezone(timezone.utc) - end).days
                else:
                    ses['DELAY'] = min((datetime.now(tz=timezone.utc) - end).days, 9999)

            sessions.append(ses)

        # Sort by date
        self.sessions = sorted(sessions, key=itemgetter('START'))
        return not self.has_errors

    def read_intensive(self, sheet: Worksheet) -> bool:
        """
        Read intensive sessions stored in Excel worksheet
        :param sheet: Active sheet
        :return: Success
        """
        rows = sheet.iter_rows()
        # Read fields from first row
        fields = self.read_header(next(rows))
        # Read data
        sessions, default_field = [], None
        for row in rows:
            ses, master = {}, []
            for cell in row:
                # Check if first cell is empty
                if isinstance(cell, EmptyCell) or cell.value == '|':
                    continue
                if 'row' not in ses:
                    ses['row'] = cell.row

                if field := fields.get(cell.column, default_field):
                    if field == 'STATIONS':
                        sta, _ = self.validate_station_info(ses, cell, field)
                        master.append(sta)
                        default_field = 'STATIONS'
                    elif field:
                        ses[field] = cell.value
                        default_field = None

            if not ses.get('DATE', None):
                continue

            # Validate this session
            ses = self.validate_session(ses, master)
            # Format station codes
            ses['master'] = self.format_list(master, 2)
            ses['DELAY'], ses['MK4NUM'] = 0, ''

            sessions.append(ses)

        # Rebuild fields
        self.fields = OrderedDict()
        for code, item in fields.items():
            if item not in self.fields.values():
                self.fields[code] = item
        self.fields[1001] = 'DELAY'
        self.fields[1002] = 'MK4NUM'

        # Sort by date
        self.sessions = sorted(sessions, key=itemgetter('START'))

        return not self.has_errors

    def process(self) -> bool:
        """
        Process an Excel file
        :return: Success
        """
        # Make sure file exists
        if not self.path.exists():
            self.exit(error=f'Could not find {self.path}')
        # Make sure it is an Excel file
        try:
            xlsx = load_workbook(self.path, read_only=True, data_only=True)
        except InvalidFileException:
            self.exit(error=f'{self.name} not a valid Excel file')

        # Use first sheet
        sheet = xlsx[xlsx.sheetnames[0]]
        # Determine file type and process it
        name = self.name.lower()
        if 'int' in name:
            self.type = 'intensives'
            return self.read_intensive(sheet)
        self.type = 'master'
        return self.read_master(sheet)

    def write_file(self, code: str) -> Path:
        """
        Write output file
        :param code: Master file code
        :return: File path
        """
        type_id = code if code in self.version else 'master'
        with MasterFile(code, self.year, self.version, self.fields) as file:
            file.write_header()
            for session in self.sessions:
                session['STATIONS'] = session[type_id]
                file.write(session)
            file.write_footer()
            return file.path

    def make_master(self) -> Path:
        """
        Make master text file
        :return: Path to output file
        """
        return self.write_file(self.type)

    def make_media(self) -> Optional[Path]:
        """
        Make media text file
        :return: Path to output file
        """
        if self.type == 'master':
            self.version['media'] = app.config['media'].get('version', self.version['master'])
            return self.write_file('media')
        return None

    def get_fs_10_stations(self) -> Tuple[int, List[str]]:
        """
        Read list of stations using FS version 10 or newer
        :return: Maximum name size for old FS version, List of stations
        """
        print(f"fs-10 {Path(self.folder, 'fs-10.toml')}")
        with open(Path(self.folder, 'fs-10.toml'), 'rb') as f:
            data = toml.loads(f.read().decode('utf-8'))
        return data['old_code_constrain'], data['fs-10']

    def get_session_type_dict(self) -> Dict[str, str]:
        """
        Read dictionary of sessions with
        :return: Dictionary of sessions with their associate type
        """
        print(f"master-type-map {Path(self.folder, 'master-type-map.json')}")
        with open(Path(self.folder, 'master-type-map.json'), 'rb') as f:
            data = json.loads(f.read().decode('utf-8'))
        return {ses_id.casefold(): ses_type for ses_type, sessions in data.items() for ses_id in sessions}




